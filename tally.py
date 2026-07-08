#!/usr/bin/env python3
"""
COMP Agent Tally — automated agent-count collector.

Replaces Emma's manual process: walks every Compass source URL already stored
in the workbook, scrapes the live "<N> Agents Found in <Location>" count off the
page, and appends a new dated column to each sheet — preserving the existing
layout so the file stays a drop-in replacement for the hand-maintained one.

Usage:
    python tally.py                       # add today's column to the default file
    python tally.py --file "COMP Agent Tally.xlsx"
    python tally.py --date 2026-07-08     # override the column date
    python tally.py --dry-run             # scrape + print, don't write the file

Only depends on openpyxl (pip install openpyxl); fetching uses the stdlib.
"""

import argparse
import datetime as dt
import gzip
import re
import shutil
import ssl
import sys
import time
import urllib.request
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter


def _ssl_context():
    """Build a verifying SSL context. python.org macOS builds ship without the
    system trust store wired up, so fall back to certifi's CA bundle."""
    try:
        import certifi
        return ssl.create_default_context(cafile=certifi.where())
    except Exception:
        try:
            return ssl.create_default_context()
        except Exception:
            return None


SSL_CTX = _ssl_context()

DEFAULT_FILE = "COMP_Agent)Tally.xlsx"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36")
COUNT_RE = re.compile(r"([\d,]+)\s+Agents?\s+Found\s+in", re.IGNORECASE)


# --------------------------------------------------------------------------- #
# Scraping
# --------------------------------------------------------------------------- #
def fetch_count(url, retries=3, pause=0.6):
    """Fetch a Compass location URL and return (count, status).

    status is one of:
      "ok"      -> count is an int
      "retired" -> the URL redirected off the location page (source removed)
      "no-data" -> page loaded but no agent count present
      "error"   -> network/HTTP failure after all retries
    """
    last_err = "error"
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": UA, "Accept-Encoding": "gzip"})
            with urllib.request.urlopen(req, timeout=30, context=SSL_CTX) as resp:
                final_url = resp.geturl()
                raw = resp.read()
                if resp.headers.get("Content-Encoding") == "gzip":
                    raw = gzip.decompress(raw)
                html = raw.decode("utf-8", "replace")
            if "/locations/" not in final_url:
                return None, "retired"      # redirected to a generic page
            m = COUNT_RE.search(html)
            if m:
                return int(m.group(1).replace(",", "")), "ok"
            last_err = "no-data"
        except Exception:  # network / decode / http error
            last_err = "error"
        time.sleep(pause * attempt)
    return None, last_err


# --------------------------------------------------------------------------- #
# Workbook helpers
# --------------------------------------------------------------------------- #
def copy_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def find_link_column(ws):
    """Column index that holds the source-URL hyperlinks, or None."""
    counts = {}
    for row in ws.iter_rows():
        for c in row:
            if c.hyperlink is not None:
                counts[c.column] = counts.get(c.column, 0) + 1
    return max(counts, key=counts.get) if counts else None


def last_date_column(ws, header_row=1, start_col=2):
    """Rightmost column in header_row whose value is a date."""
    last = start_col - 1
    for c in range(start_col, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, (dt.date, dt.datetime)):
            last = c
    return last


def shift_column_right(ws, col):
    """Move an entire column one to the right (value + hyperlink + style)."""
    for r in range(1, ws.max_row + 1):
        src = ws.cell(r, col)
        dst = ws.cell(r, col + 1)
        dst.value = src.value
        copy_style(src, dst)
        if src.hyperlink is not None:
            dst.hyperlink = copy(src.hyperlink)
            dst.hyperlink.ref = dst.coordinate
        src.value = None
        src.hyperlink = None


def existing_date_column(ws, header_row, target_date):
    if isinstance(target_date, dt.datetime):
        target_date = target_date.date()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, dt.datetime):
            v = v.date()
        if isinstance(v, dt.date) and v == target_date:
            return c
    return None


# --------------------------------------------------------------------------- #
# Sheet processors
# --------------------------------------------------------------------------- #
def process_by_state(ws, run_date, dry_run):
    """By State: dates in row 1, 'Agent Count' in row 2, state in col A,
    source hyperlink in the link column. Append a new dated column just left
    of the link column, pushing the link column right."""
    link_col = find_link_column(ws)
    if link_col is None:
        print("  (no source links found — skipping)")
        return []

    header_row, label_row = 1, 2

    # The workbook parks an excluded "District of Columbia" link below the
    # Total row; only the contiguous state block above Total is real data.
    total_row = next((r for r in range(1, ws.max_row + 1)
                      if str(ws.cell(r, 1).value).strip().lower() == "total"),
                     ws.max_row + 1)
    data_rows = [r for r in range(1, total_row)
                 if ws.cell(r, link_col).hyperlink is not None]
    if not data_rows:
        return []
    first_data, last_data = min(data_rows), max(data_rows)

    # scrape first, using the URLs currently in place
    results = []
    for r in data_rows:
        state = ws.cell(r, 1).value
        url = ws.cell(r, link_col).hyperlink.target
        count, status = fetch_count(url)
        results.append((state, count))
        flag = "" if status == "ok" else f"  <{status}>"
        print(f"    {str(state):>4}  {('' if count is None else count):>6}{flag}")

    if dry_run:
        return results

    # decide target column (overwrite if we already ran today, else insert)
    new_col = existing_date_column(ws, header_row, run_date)
    if new_col is None:
        new_col = last_date_column(ws, header_row) + 1
        if new_col >= link_col:          # need room before the link column
            shift_column_right(ws, link_col)
            link_col += 1

    # sample styling from the previous date column
    prev = new_col - 1
    hdr_c = ws.cell(header_row, new_col)
    hdr_c.value = run_date
    copy_style(ws.cell(header_row, prev), hdr_c)
    if not isinstance(ws.cell(header_row, prev).value, (dt.date, dt.datetime)):
        hdr_c.number_format = "mm-dd-yy"
    lbl_c = ws.cell(label_row, new_col)
    lbl_c.value = "Agent Count"
    copy_style(ws.cell(label_row, prev), lbl_c)

    by_row = {r: res for r, res in zip(data_rows, results)}
    for r in data_rows:
        _, count = by_row[r]
        cell = ws.cell(r, new_col)
        cell.value = count
        copy_style(ws.cell(r, prev), cell)

    # keep a SUM in the Total row for the new column, if that row exists
    col_letter = get_column_letter(new_col)
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip().lower() == "total":
            tc = ws.cell(r, new_col)
            tc.value = f"=SUM({col_letter}{first_data}:{col_letter}{last_data})"
            copy_style(ws.cell(r, prev), tc)
            break
    return results


def process_by_market(ws, run_date, dry_run):
    """By Market: 'Source URL' in A1, one URL per row below. Track counts in
    dated columns to the right (date in row 1, counts aligned to each URL)."""
    link_col = find_link_column(ws)
    if link_col is None or link_col != 1:
        return []
    url_rows = [r for r in range(2, ws.max_row + 1)
                if ws.cell(r, link_col).hyperlink is not None]
    if not url_rows:
        return []

    results = []
    for r in url_rows:
        url = ws.cell(r, link_col).hyperlink.target
        count, status = fetch_count(url)
        slug = url.rstrip("/").split("/")[-2]
        results.append((slug, count))
        flag = "" if status == "ok" else f"  <{status}>"
        print(f"    {slug:>22}  {('' if count is None else count):>6}{flag}")

    if dry_run:
        return results

    new_col = existing_date_column(ws, 1, run_date)
    if new_col is None:
        new_col = max(last_date_column(ws, 1, start_col=2), 1) + 1
    ws.cell(1, new_col).value = run_date
    ws.cell(1, new_col).number_format = "mm-dd-yy"
    for r, (_, count) in zip(url_rows, results):
        ws.cell(r, new_col).value = count
    return results


# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="Append a dated agent-count column.")
    ap.add_argument("--file", default=DEFAULT_FILE)
    ap.add_argument("--date", help="YYYY-MM-DD column date (default: today)")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--no-backup", action="store_true")
    args = ap.parse_args()

    run_date = (dt.datetime.strptime(args.date, "%Y-%m-%d").date()
                if args.date else dt.date.today())
    run_dt = dt.datetime(run_date.year, run_date.month, run_date.day)

    if not args.dry_run and not args.no_backup:
        backup = f"{args.file}.{run_date:%Y%m%d}.bak"
        shutil.copy2(args.file, backup)
        print(f"Backup: {backup}")

    wb = openpyxl.load_workbook(args.file)
    print(f"\nRun date: {run_date}   File: {args.file}")

    for name, fn in (("By State", process_by_state),
                     ("By Market", process_by_market)):
        if name in wb.sheetnames:
            print(f"\n[{name}]")
            res = fn(wb[name], run_dt, args.dry_run)
            got = [c for _, c in res if c is not None]
            print(f"  -> {len(got)}/{len(res)} scraped"
                  + (f", total = {sum(got):,}" if got else ""))

    if args.dry_run:
        print("\nDry run — file not modified.")
    else:
        wb.save(args.file)
        print(f"\nSaved column {run_date} to {args.file}")


if __name__ == "__main__":
    main()
