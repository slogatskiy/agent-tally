#!/usr/bin/env python3
"""
Build the public HTML widget from COMP_Agent)Tally.xlsx.

Reads both sheets, extracts the full dated time series, writes:
  - data/history.json   (raw series, for reuse / transparency)
  - docs/index.html      (self-contained dashboard for GitHub Pages)

The page embeds its own data, so it renders offline and needs no network.
Charts follow the validated data-viz palette (single blue series, theme-aware).
"""

import datetime as dt
import json
import html
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "COMP_Agent)Tally.xlsx"
DATA = ROOT / "data" / "history.json"
SITE = ROOT / "docs" / "index.html"

STATE_NAMES = {
    "CA": "California", "CO": "Colorado", "FL": "Florida", "HI": "Hawaii",
    "IL": "Illinois", "KS": "Kansas", "MA": "Massachusetts", "MD": "Maryland",
    "MO": "Missouri", "NC": "North Carolina", "NJ": "New Jersey", "NV": "Nevada",
    "NY": "New York", "PA": "Pennsylvania", "TX": "Texas", "VA": "Virginia",
    "WA": "Washington", "WI": "Wisconsin", "WY": "Wyoming", "OR": "Oregon",
    "ID": "Idaho", "AZ": "Arizona", "TN": "Tennessee", "MN": "Minnesota",
    "LA": "Louisiana", "MS": "Mississippi", "IN": "Indiana", "GA": "Georgia",
    "SC": "South Carolina", "DE": "Delaware", "CT": "Connecticut",
    "NH": "New Hampshire", "ME": "Maine",
}


def _date_cols(ws, header_row=1, start_col=2):
    cols = []
    for c in range(start_col, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, dt.datetime):
            cols.append((c, v.date()))
        elif isinstance(v, dt.date):
            cols.append((c, v))
    return cols


def _num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def extract_by_state(ws):
    date_cols = _date_cols(ws)
    dates = [d.isoformat() for _, d in date_cols]
    link_col = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink:
                link_col = cell.column
                break
        if link_col:
            break
    total_row = next((r for r in range(1, ws.max_row + 1)
                      if str(ws.cell(r, 1).value).strip().lower() == "total"),
                     ws.max_row + 1)
    rows = []
    for r in range(1, total_row):
        code = ws.cell(r, 1).value
        cell = ws.cell(r, link_col) if link_col else None
        if not (cell and cell.hyperlink):
            continue
        counts = [_num(ws.cell(r, c).value) for c, _ in date_cols]
        rows.append({
            "code": code,
            "name": STATE_NAMES.get(code, code),
            "counts": counts,
            "url": cell.hyperlink.target,
        })
    totals = []
    for i in range(len(dates)):
        vals = [row["counts"][i] for row in rows if row["counts"][i] is not None]
        totals.append(sum(vals) if vals else None)
    return {"dates": dates, "rows": rows, "totals": totals}


def extract_by_market(ws):
    date_cols = _date_cols(ws)
    dates = [d.isoformat() for _, d in date_cols]
    rows = []
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, 1)
        if not cell.hyperlink:
            continue
        url = cell.hyperlink.target
        slug = url.rstrip("/").split("/")[-2]
        name = slug.rsplit("-", 1)[0].replace("-", " ").title()
        counts = [_num(ws.cell(r, c).value) for c, _ in date_cols]
        rows.append({"code": slug, "name": name, "counts": counts, "url": url})
    totals = []
    for i in range(len(dates)):
        vals = [row["counts"][i] for row in rows if row["counts"][i] is not None]
        totals.append(sum(vals) if vals else None)
    return {"dates": dates, "rows": rows, "totals": totals}


def build_data():
    wb = openpyxl.load_workbook(XLSX)
    data = {
        "generated": dt.date.today().isoformat(),
        "source": "compass.com/agents/locations",
        "by_state": extract_by_state(wb["By State"]),
        "by_market": extract_by_market(wb["By Market"]),
    }
    DATA.parent.mkdir(parents=True, exist_ok=True)
    DATA.write_text(json.dumps(data, indent=2))
    return data


# --------------------------------------------------------------------------- #
# HTML rendering
# --------------------------------------------------------------------------- #
def fmt(n):
    return "—" if n is None else f"{n:,}"


def fmt_date(iso):
    d = dt.date.fromisoformat(iso)
    return d.strftime("%b %-d, %Y")


def delta_cell(cur, prev):
    if cur is None or prev is None:
        return '<td class="delta"></td>'
    d = cur - prev
    if d == 0:
        return '<td class="delta flat">0</td>'
    cls = "up" if d > 0 else "down"
    arrow = "▲" if d > 0 else "▼"
    return f'<td class="delta {cls}">{arrow} {d:+,}</td>'


def render_table(section, latest_only_delta=True):
    dates = section["dates"]
    head = "".join(f"<th class='num'>{fmt_date(d)}</th>" for d in dates)
    body = []
    for row in sorted(section["rows"], key=lambda x: -(x["counts"][-1] or -1)):
        counts = row["counts"]
        tds = "".join(f"<td class='num'>{fmt(c)}</td>" for c in counts)
        prev = counts[-2] if len(counts) >= 2 else None
        dcell = delta_cell(counts[-1], prev)
        name = html.escape(row["name"])
        code = html.escape(str(row["code"]))
        label = (f"<a href='{html.escape(row['url'])}' target='_blank' "
                 f"rel='noopener'>{name}</a>")
        spark = (f"<td class='spark'><svg data-spark='{code}' "
                 f"width='72' height='22'></svg></td>")
        body.append(
            f"<tr data-code='{code}'><th class='rowh'>{label}</th>"
            f"{spark}{tds}{dcell}</tr>")
    tot = section["totals"]
    tot_tds = "".join(f"<td class='num'>{fmt(c)}</td>" for c in tot)
    tot_prev = tot[-2] if len(tot) >= 2 else None
    tot_delta = delta_cell(tot[-1], tot_prev)
    total_row = (f"<tr class='total'><th class='rowh'>Total</th>"
                 f"<td class='spark'></td>{tot_tds}{tot_delta}</tr>")
    return f"""
    <table>
      <thead><tr><th class='rowh'>Location</th><th>Trend</th>{head}
        <th class='num'>MoM</th></tr></thead>
      <tbody>{''.join(body)}{total_row}</tbody>
    </table>"""


def render(data):
    st, mk = data["by_state"], data["by_market"]
    st_tot = [t for t in st["totals"] if t is not None]
    latest = st_tot[-1] if st_tot else 0
    prev = st_tot[-2] if len(st_tot) >= 2 else None
    if prev:
        d = latest - prev
        pct = d / prev * 100
        arrow = "▲" if d > 0 else ("▼" if d < 0 else "—")
        dcls = "up" if d > 0 else ("down" if d < 0 else "flat")
        hero_delta = (f"<span class='hero-delta {dcls}'>{arrow} {d:+,} "
                      f"({pct:+.1f}%) vs prior</span>")
    else:
        hero_delta = ""
    n_states = len([r for r in st["rows"] if r["counts"][-1] is not None])
    mk_live = len([r for r in mk["rows"] if r["counts"][-1] is not None])
    updated = fmt_date(st["dates"][-1]) if st["dates"] else "—"

    payload = json.dumps({
        "stateDates": st["dates"], "stateTotals": st["totals"],
        "stateRows": [{"code": r["code"], "counts": r["counts"]}
                      for r in st["rows"]],
        "marketRows": [{"code": r["code"], "counts": r["counts"]}
                       for r in mk["rows"]],
    })

    return TEMPLATE.format(
        hero=fmt(latest), hero_delta=hero_delta, n_states=n_states,
        mk_live=mk_live, updated=html.escape(updated),
        generated=html.escape(fmt_date(data["generated"])),
        state_table=render_table(st), market_table=render_table(mk),
        payload=payload)


TEMPLATE = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Compass Agent Tally</title>
<style>
  :root {{
    --plane:#f9f9f7; --surface:#fcfcfb; --ink:#0b0b0b; --ink2:#52514e;
    --muted:#898781; --grid:#e1e0d9; --axis:#c3c2b7; --series:#2a78d6;
    --up:#006300; --down:#c0392b; --border:rgba(11,11,11,0.10);
  }}
  @media (prefers-color-scheme: dark) {{
    :root {{
      --plane:#0d0d0d; --surface:#1a1a19; --ink:#fff; --ink2:#c3c2b7;
      --muted:#898781; --grid:#2c2c2a; --axis:#383835; --series:#3987e5;
      --up:#0ca30c; --down:#e66767; --border:rgba(255,255,255,0.10);
    }}
  }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; background:var(--plane); color:var(--ink);
    font:15px/1.5 system-ui,-apple-system,"Segoe UI",sans-serif; }}
  .wrap {{ max-width:1120px; margin:0 auto; padding:32px 20px 64px; }}
  header h1 {{ font-size:20px; margin:0 0 2px; letter-spacing:-.01em; }}
  header p {{ margin:0; color:var(--ink2); font-size:13px; }}
  .cards {{ display:flex; flex-wrap:wrap; gap:16px; margin:24px 0 8px; }}
  .card {{ background:var(--surface); border:1px solid var(--border);
    border-radius:12px; padding:18px 20px; flex:1; min-width:180px; }}
  .card .label {{ font-size:12px; color:var(--muted); text-transform:uppercase;
    letter-spacing:.06em; }}
  .card .value {{ font-size:32px; font-weight:650; margin-top:6px;
    letter-spacing:-.02em; }}
  .hero-delta {{ font-size:14px; font-weight:550; }}
  .up {{ color:var(--up); }} .down {{ color:var(--down); }}
  .flat {{ color:var(--muted); }}
  .chart-card {{ background:var(--surface); border:1px solid var(--border);
    border-radius:12px; padding:20px; margin:16px 0 8px; }}
  .chart-card h2, section h2 {{ font-size:14px; margin:0 0 4px;
    color:var(--ink2); font-weight:600; }}
  .chart-card .sub {{ font-size:12px; color:var(--muted); margin:0 0 12px; }}
  section {{ margin-top:32px; }}
  .scroll {{ overflow-x:auto; border:1px solid var(--border);
    border-radius:12px; background:var(--surface); }}
  table {{ border-collapse:collapse; width:100%; font-size:13px; }}
  th,td {{ padding:7px 12px; text-align:left; white-space:nowrap; }}
  thead th {{ position:sticky; top:0; background:var(--surface);
    color:var(--muted); font-weight:600; font-size:11px;
    text-transform:uppercase; letter-spacing:.05em;
    border-bottom:1px solid var(--axis); }}
  tbody tr {{ border-bottom:1px solid var(--grid); }}
  tbody tr:hover {{ background:rgba(42,120,214,0.06); }}
  .rowh {{ font-weight:500; }}
  .rowh a {{ color:var(--ink); text-decoration:none;
    border-bottom:1px solid transparent; }}
  .rowh a:hover {{ border-bottom-color:var(--series); }}
  .num {{ text-align:right; font-variant-numeric:tabular-nums; }}
  .delta {{ text-align:right; font-variant-numeric:tabular-nums;
    font-weight:550; font-size:12px; }}
  .spark {{ padding:4px 8px; }}
  tr.total {{ font-weight:700; border-top:2px solid var(--axis); }}
  tr.total td, tr.total th {{ background:var(--surface); }}
  .tabs {{ display:flex; gap:6px; margin-bottom:12px; }}
  .tab {{ padding:6px 14px; border:1px solid var(--border);
    border-radius:999px; background:transparent; color:var(--ink2);
    cursor:pointer; font:inherit; font-size:13px; }}
  .tab.active {{ background:var(--series); color:#fff; border-color:var(--series); }}
  footer {{ margin-top:40px; color:var(--muted); font-size:12px; }}
  footer a {{ color:var(--muted); }}
  .hidden {{ display:none; }}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>Compass Agent Tally</h1>
    <p>Compass, Inc. (COMP) self-reported real-estate agent headcount by state and
       market · updated {updated}</p>
  </header>

  <div class="cards">
    <div class="card">
      <div class="label">Total agents (states)</div>
      <div class="value">{hero}</div>
      <div>{hero_delta}</div>
    </div>
    <div class="card">
      <div class="label">States tracked</div>
      <div class="value">{n_states}</div>
    </div>
    <div class="card">
      <div class="label">Markets tracked</div>
      <div class="value">{mk_live}</div>
    </div>
  </div>

  <div class="chart-card">
    <h2>Total agents over time</h2>
    <p class="sub">Sum across all tracked states, per collection date.</p>
    <svg id="trend" width="100%" height="220" preserveAspectRatio="none"></svg>
  </div>

  <section>
    <div class="tabs">
      <button class="tab active" data-tab="state">By State</button>
      <button class="tab" data-tab="market">By Market</button>
    </div>
    <div id="tab-state">
      <div class="scroll">{state_table}</div>
    </div>
    <div id="tab-market" class="hidden">
      <div class="scroll">{market_table}</div>
    </div>
  </section>

  <footer>
    Data scraped from <a href="https://www.compass.com/agents/locations/"
    target="_blank" rel="noopener">compass.com</a> · figures are Compass's own
    counts · generated {generated} · auto-updated monthly.
  </footer>
</div>

<script id="data" type="application/json">{payload}</script>
<script>
const D = JSON.parse(document.getElementById('data').textContent);
const css = k => getComputedStyle(document.body).getPropertyValue(k).trim();

function sparkline(svg, counts) {{
  const pts = counts.map((v,i)=>[i,v]).filter(p=>p[1]!=null);
  if (pts.length < 2) return;
  const w=+svg.getAttribute('width'), h=+svg.getAttribute('height'), pad=3;
  const xs=pts.map(p=>p[0]), ys=pts.map(p=>p[1]);
  const x0=Math.min(...xs), x1=Math.max(...xs);
  const y0=Math.min(...ys), y1=Math.max(...ys);
  const sx=i=>pad+(x1===x0?0:(i-x0)/(x1-x0))*(w-2*pad);
  const sy=v=>h-pad-(y1===y0?(h-2*pad)/2:(v-y0)/(y1-y0)*(h-2*pad));
  const d=pts.map((p,i)=>(i?'L':'M')+sx(p[0]).toFixed(1)+' '+sy(p[1]).toFixed(1)).join(' ');
  const up = pts[pts.length-1][1] >= pts[0][1];
  const col = up ? css('--up') : css('--down');
  svg.innerHTML =
    `<path d="${{d}}" fill="none" stroke="${{col}}" stroke-width="1.5"/>`+
    `<circle cx="${{sx(pts[pts.length-1][0]).toFixed(1)}}" `+
    `cy="${{sy(pts[pts.length-1][1]).toFixed(1)}}" r="2" fill="${{col}}"/>`;
}}
document.querySelectorAll('svg[data-spark]').forEach(svg=>{{
  const code=svg.getAttribute('data-spark');
  const row=[...D.stateRows,...D.marketRows].find(r=>String(r.code)===code);
  if(row) sparkline(svg,row.counts);
}});

function trend() {{
  const svg=document.getElementById('trend');
  const W=svg.clientWidth||900, H=220, L=54,R=16,T=16,B=34;
  const dates=D.stateDates, vals=D.stateTotals;
  const idx=vals.map((v,i)=>[i,v]).filter(p=>p[1]!=null);
  if(idx.length<2){{svg.innerHTML='';return;}}
  const ys=idx.map(p=>p[1]);
  let lo=Math.min(...ys), hi=Math.max(...ys); const padv=(hi-lo)*0.15||1;
  lo-=padv; hi+=padv;
  const sx=i=>L+i/(dates.length-1)*(W-L-R);
  const sy=v=>T+(1-(v-lo)/(hi-lo))*(H-T-B);
  const ink2=css('--ink2'), muted=css('--muted'), grid=css('--grid'),
        series=css('--series'), surf=css('--surface');
  let g='';
  const ticks=4;
  for(let t=0;t<=ticks;t++){{
    const val=lo+(hi-lo)*t/ticks, y=sy(val);
    g+=`<line x1="${{L}}" y1="${{y.toFixed(1)}}" x2="${{W-R}}" y2="${{y.toFixed(1)}}" stroke="${{grid}}" stroke-width="1"/>`;
    g+=`<text x="${{L-8}}" y="${{(y+3).toFixed(1)}}" text-anchor="end" font-size="11" fill="${{muted}}">${{Math.round(val).toLocaleString()}}</text>`;
  }}
  dates.forEach((dd,i)=>{{
    const lab=new Date(dd+'T00:00').toLocaleDateString('en-US',{{month:'short',day:'numeric'}});
    g+=`<text x="${{sx(i).toFixed(1)}}" y="${{H-12}}" text-anchor="middle" font-size="11" fill="${{muted}}">${{lab}}</text>`;
  }});
  const dpath=idx.map((p,i)=>(i?'L':'M')+sx(p[0]).toFixed(1)+' '+sy(p[1]).toFixed(1)).join(' ');
  g+=`<path d="${{dpath}}" fill="none" stroke="${{series}}" stroke-width="2"/>`;
  idx.forEach(p=>{{
    g+=`<circle cx="${{sx(p[0]).toFixed(1)}}" cy="${{sy(p[1]).toFixed(1)}}" r="3.5" fill="${{series}}" stroke="${{surf}}" stroke-width="2"/>`;
  }});
  const last=idx[idx.length-1];
  g+=`<text x="${{(sx(last[0])-6).toFixed(1)}}" y="${{(sy(last[1])-9).toFixed(1)}}" text-anchor="end" font-size="12" font-weight="600" fill="${{ink2}}">${{last[1].toLocaleString()}}</text>`;
  svg.innerHTML=g;
}}
trend();
addEventListener('resize',trend);

document.querySelectorAll('.tab').forEach(btn=>btn.onclick=()=>{{
  document.querySelectorAll('.tab').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  const t=btn.dataset.tab;
  document.getElementById('tab-state').classList.toggle('hidden',t!=='state');
  document.getElementById('tab-market').classList.toggle('hidden',t!=='market');
}});
</script>
</body>
</html>"""


def main():
    data = build_data()
    SITE.parent.mkdir(parents=True, exist_ok=True)
    SITE.write_text(render(data))
    st = data["by_state"]["totals"]
    print(f"Built {SITE.relative_to(ROOT)} — {len(data['by_state']['dates'])} dates,"
          f" latest state total {fmt(st[-1] if st else None)}")


if __name__ == "__main__":
    main()
